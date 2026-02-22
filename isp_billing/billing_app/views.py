from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.forms import (
    modelformset_factory,
    ModelChoiceField,
    ModelMultipleChoiceField
)
from django.contrib.auth.decorators import (
    login_required as django_login_required
)
from django.contrib.auth import logout
from django.views.decorators.http import require_http_methods
from django.views.decorators.cache import never_cache
from datetime import datetime, date, time
from io import BytesIO
from decimal import Decimal, InvalidOperation
import logging
import random
from typing import Any, cast
from .models import (
    Cliente, Distrito, Sector, Via, Plan, ClientePlan, Pago,
    SerieCorrelativo, Servicio, OrdenTecnicaConcepto, OrdenTecnica,
    MikrotikConfig, Tecnico, DeudaExcluida, EgresoConcepto, Egreso,
    AppRole, UserRole, CompanySettings
)
from django.contrib.auth import get_user_model
from .forms import (
    ClienteForm, ClientePlanForm, OrdenTecnicaForm,
    DistritoForm, SectorForm, ViaForm,
    ServicioForm, PlanForm, OTConceptoForm, SerieCorrelativoForm,
    MikrotikConfigForm, TecnicoForm, EgresoConceptoForm, EgresoForm,
    AppRoleForm, UserRoleForm, UserCreateForm, UserEditForm,
    CompanySettingsForm
)
from .utils import calcular_meses_deuda, registrar_movimiento
from .permissions import (
    can_delete_cliente,
    can_manage_ajustes,
    can_manage_caja,
    can_cobrar,
    is_developer,
)

logger = logging.getLogger(__name__)


def login_required(*args, **kwargs):
    kwargs['login_url'] = 'login'
    return django_login_required(*args, **kwargs)


@never_cache
def logout_view(request):
    """Vista personalizada de logout que cierra sesión completamente"""
    logout(request)
    response = redirect('/login/')
    # Evitar caché del navegador
    response['Cache-Control'] = (
        'no-cache, no-store, must-revalidate, '
        'post-check=0, pre-check=0'
    )
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def _cliente_planes_qs(cliente):
    """Devuelve queryset de planes del cliente (relación dinámica)."""
    return cast(Any, cliente).planes


def _cliente_servicios_qs(cliente):
    return Servicio.objects.filter(
        planes__clienteplan__cliente=cliente
    ).distinct().order_by('nombre')


def _to_aware_datetime(value):
    """Convierte date/datetime a datetime aware para ordenar con seguridad."""
    if value is None:
        return None
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return value
        return timezone.make_aware(value)
    if isinstance(value, date):
        return timezone.make_aware(datetime.combine(value, time.min))
    return None


def _to_decimal(value):
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _normalize_excel_value(value):
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return timezone.localtime(value).replace(tzinfo=None)
        return value
    return value


def _clean_text(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _normalize_digits(value, length):
    if value is None:
        return ''
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return str(int(value)).zfill(length)
        except (TypeError, ValueError):
            return ''
    text = _clean_text(value).replace(' ', '').replace('-', '')
    if text.isdigit():
        return text.zfill(length)
    return text


def _parse_import_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean_text(value)
    parsed = parse_date(text)
    if parsed:
        return parsed
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _get_or_create_distrito(nombre):
    distrito = Distrito.objects.filter(nombre__iexact=nombre).first()
    if distrito:
        return distrito
    return Distrito.objects.create(nombre=nombre)


def _get_or_create_sector(distrito, nombre):
    sector = Sector.objects.filter(
        distrito=distrito, nombre__iexact=nombre
    ).first()
    if sector:
        return sector
    return Sector.objects.create(distrito=distrito, nombre=nombre)


def _get_or_create_via(sector, nombre):
    via = Via.objects.filter(
        sector=sector, nombre__iexact=nombre, tipo='CALLE'
    ).first()
    if via:
        return via
    return Via.objects.create(sector=sector, nombre=nombre, tipo='CALLE')


def _import_clientes_con_calles_xlsx(file_obj):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            'import_error': 'Falta instalar openpyxl para leer XLSX.'
        }

    wb = load_workbook(file_obj, data_only=True)
    ws = wb.active
    if ws is None:
        return {
            'import_error': 'No se pudo leer la hoja activa del archivo.'
        }
    summary = {
        'total': 0,
        'created': 0,
        'skipped': 0,
        'errors': 0,
    }
    errors = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row or not any(row):
            continue
        summary['total'] += 1

        dni = _normalize_digits(row[0], 8)
        apellidos = _clean_text(row[1])
        nombres = _clean_text(row[2])
        celular = _normalize_digits(row[3], 9)
        distrito_name = _clean_text(row[4])
        sector_name = _clean_text(row[5])
        via_name = _clean_text(row[6])
        estado_raw = _clean_text(row[7]).lower()
        fecha_raw = row[8] if len(row) > 8 else None
        fecha_instalacion = _parse_import_date(fecha_raw)

        row_errors = []
        if not dni or len(dni) != 8:
            row_errors.append('DNI invalido')
        if not apellidos:
            row_errors.append('Apellidos requeridos')
        if not nombres:
            row_errors.append('Nombres requeridos')
        if not celular or len(celular) != 9:
            row_errors.append('Celular invalido')
        if not distrito_name:
            row_errors.append('Distrito requerido')
        if not sector_name:
            row_errors.append('Sector requerido')
        if not via_name:
            row_errors.append('Via requerida')
        if fecha_raw not in (None, '') and not fecha_instalacion:
            row_errors.append('Fecha de instalacion invalida')

        if row_errors:
            summary['errors'] += 1
            errors.append({
                'row': row_idx,
                'errors': row_errors,
            })
            continue

        if Cliente.objects.filter(dni=dni).exists():
            summary['skipped'] += 1
            continue

        distrito = _get_or_create_distrito(distrito_name)
        sector = _get_or_create_sector(distrito, sector_name)
        via = _get_or_create_via(sector, via_name)
        estado_activo = estado_raw in (
            'activo', '1', 'true', 'si', 'si.'
        )
        if estado_raw == 'inactivo':
            estado_activo = False
        if estado_raw == '':
            estado_activo = True

        cliente = Cliente(
            dni=dni,
            apellidos=apellidos,
            nombres=nombres,
            celular=celular,
            via=via,
            estado_activo=estado_activo,
        )
        if fecha_instalacion:
            cliente.fecha_instalacion = fecha_instalacion
        cliente.save()
        summary['created'] += 1

    return {
        'import_summary': summary,
        'import_errors': errors,
    }


def _write_excel_sheet(ws, headers, rows):
    from openpyxl.utils import get_column_letter

    ws.append(headers)
    for row in rows:
        ws.append([_normalize_excel_value(cell) for cell in row])

    for idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = (
            min(max_len + 2, 50) if max_len else 12
        )


def _excel_response(filename, sheets):
    from openpyxl import Workbook

    wb = Workbook()
    first = sheets[0]
    ws = cast(Any, wb.active)
    ws.title = first['title']
    _write_excel_sheet(ws, first['headers'], first['rows'])
    for sheet in sheets[1:]:
        ws = wb.create_sheet(title=sheet['title'])
        _write_excel_sheet(ws, sheet['headers'], sheet['rows'])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{filename}"'
    )
    return response


@login_required(login_url='admin:login')
def dashboard(request):
    """Dashboard principal con estadísticas"""
    hoy = timezone.localdate()
    week_start = hoy - timezone.timedelta(days=hoy.weekday())
    month_start = hoy.replace(day=1)

    pagos_hoy = Pago.objects.filter(
        fecha__date=hoy
    ).aggregate(Sum('monto'))['monto__sum'] or 0
    pagos_semana = (
        Pago.objects.filter(
            fecha__date__gte=week_start,
            fecha__date__lte=hoy
        ).aggregate(Sum('monto'))['monto__sum'] or 0
    )
    pagos_mes = (
        Pago.objects.filter(
            fecha__date__gte=month_start,
            fecha__date__lte=hoy
        ).aggregate(Sum('monto'))['monto__sum'] or 0
    )
    egresos_mes = (
        Egreso.objects.filter(
            fecha__gte=month_start,
            fecha__lte=hoy
        ).aggregate(Sum('monto'))['monto__sum'] or 0
    )
    balance_mes = _to_decimal(pagos_mes) - _to_decimal(egresos_mes)

    ots_pendientes = OrdenTecnica.objects.filter(completada=False).count()
    ots_completadas = OrdenTecnica.objects.filter(completada=True).count()
    ots_pagadas = OrdenTecnica.objects.filter(pagada=True).count()

    context = {
        'total_clientes': Cliente.objects.count(),
        'clientes_activos': Cliente.objects.filter(estado_activo=True).count(),
        'total_planes': ClientePlan.objects.filter(activo=True).count(),
        'ots_pendientes': ots_pendientes,
        'ots_completadas': ots_completadas,
        'ots_pagadas': ots_pagadas,
        'pagos_hoy': pagos_hoy,
        'pagos_semana': pagos_semana,
        'pagos_mes': pagos_mes,
        'egresos_mes': egresos_mes,
        'balance_mes': balance_mes,
        'ultimos_clientes': (
            Cliente.objects.all().order_by('-fecha_registro')[:5]
        ),
        'ultimos_pagos': (
            Pago.objects.select_related('cliente')
            .order_by('-fecha')[:5]
        ),
        'ultimas_ots': (
            OrdenTecnica.objects.select_related('cliente', 'concepto')
            .order_by('-fecha_creacion')[:5]
        ),
    }
    return render(request, 'billing_app/dashboard.html', context)


@login_required(login_url='admin:login')
def caja_dashboard(request):
    if not can_manage_caja(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    hoy = timezone.localdate()
    periodo = (request.GET.get('periodo') or 'dia').lower()
    fecha_inicio_raw = request.GET.get('fecha_inicio')
    fecha_fin_raw = request.GET.get('fecha_fin')
    start_date = None
    end_date = None

    if fecha_inicio_raw or fecha_fin_raw:
        try:
            if fecha_inicio_raw:
                start_date = datetime.strptime(
                    fecha_inicio_raw, '%Y-%m-%d'
                ).date()
            if fecha_fin_raw:
                end_date = datetime.strptime(
                    fecha_fin_raw, '%Y-%m-%d'
                ).date()
        except (ValueError, TypeError):
            start_date = None
            end_date = None

    if start_date or end_date:
        if not start_date:
            start_date = end_date or hoy
        if not end_date:
            end_date = start_date
        periodo = 'rango'
    else:
        end_date = hoy
        if periodo == 'semanal':
            start_date = hoy - timezone.timedelta(days=6)
        elif periodo == 'quincenal':
            start_date = hoy - timezone.timedelta(days=14)
        elif periodo == 'mensual':
            start_date = hoy.replace(day=1)
        elif periodo == 'trimestral':
            quarter_start_month = ((hoy.month - 1) // 3) * 3 + 1
            start_date = hoy.replace(month=quarter_start_month, day=1)
        elif periodo == 'anual':
            start_date = hoy.replace(month=1, day=1)
        else:
            periodo = 'dia'
            start_date = hoy

    ingresos_qs = (
        Pago.objects.filter(
            fecha__date__gte=start_date,
            fecha__date__lte=end_date
        )
        .select_related('cliente')
        .order_by('-fecha')
    )
    ingresos_total = (
        ingresos_qs.aggregate(total=Sum('monto'))['total'] or 0
    )
    egresos_qs = (
        Egreso.objects.filter(fecha__gte=start_date, fecha__lte=end_date)
        .select_related('concepto')
        .order_by('-created_at')
    )
    total_egresos = (
        egresos_qs.aggregate(total=Sum('monto'))['total'] or 0
    )
    balance = _to_decimal(ingresos_total) - _to_decimal(total_egresos)

    if request.method == 'POST':
        form = EgresoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('caja-dashboard')
    else:
        form = EgresoForm(initial={'fecha': hoy})

    return render(
        request,
        'billing_app/caja/dashboard.html',
        {
            'periodo': periodo,
            'start_date': start_date,
            'end_date': end_date,
            'fecha_inicio': fecha_inicio_raw,
            'fecha_fin': fecha_fin_raw,
            'ingresos_total': ingresos_total,
            'ingresos_list': ingresos_qs,
            'egresos_hoy': egresos_qs,
            'total_egresos': total_egresos,
            'balance_hoy': balance,
            'form_egreso': form,
        }
    )


@login_required(login_url='admin:login')
def pagos_lista(request):
    if not (can_manage_caja(request.user) or can_cobrar(request.user)):
        return HttpResponseForbidden('Acceso no autorizado')

    query = (request.GET.get('q') or '').strip()
    pagos = (
        Pago.objects.select_related('cliente')
        .order_by('-fecha')
    )
    if query:
        pagos = pagos.filter(
            Q(cliente__nombres__icontains=query)
            | Q(cliente__apellidos__icontains=query)
            | Q(cliente__dni__icontains=query)
            | Q(serie_numero__icontains=query)
            | Q(tipo_comprobante__icontains=query)
        )

    return render(
        request,
        'billing_app/pagos/lista.html',
        {
            'pagos': pagos,
            'query': query,
        }
    )


@login_required(login_url='admin:login')
def reportes_index(request):
    ingresos_total = (
        Pago.objects.aggregate(total=Sum('monto'))['total'] or 0
    )
    egresos_total = (
        Egreso.objects.aggregate(total=Sum('monto'))['total'] or 0
    )
    total_clientes = Cliente.objects.count()
    total_planes = ClientePlan.objects.count()
    total_servicios = Servicio.objects.count()
    total_nodos = MikrotikConfig.objects.count()

    instalaciones = OrdenTecnica.objects.filter(
        concepto__categoria='INSTALACION'
    ).count()
    cortes = OrdenTecnica.objects.filter(
        concepto__categoria='CORTES'
    ).count()
    averias = OrdenTecnica.objects.filter(
        concepto__categoria='AVERIAS'
    ).count()
    ots_completadas = OrdenTecnica.objects.filter(completada=True).count()

    clientes_por_zona = (
        Distrito.objects.annotate(total=Count('sectores__vias__clientes'))
        .order_by('-total')
    )
    zonas_por_sector = (
        Sector.objects.annotate(total=Count('vias__clientes'))
        .order_by('-total')
    )

    deuda_rows = []
    deuda_total = Decimal('0')
    for cliente in Cliente.objects.all():
        cliente_id = cast(Any, cliente).id
        try:
            deuda = calcular_meses_deuda(cliente)
        except Exception:
            logger.exception(
                'Error calculando deuda para cliente %s',
                cliente_id
            )
            deuda = []
        total_cliente = sum(
            (_to_decimal(item.get('saldo')) for item in deuda),
            Decimal('0')
        )
        if total_cliente > 0:
            deuda_rows.append({
                'cliente': cliente,
                'total': total_cliente
            })
        deuda_total += total_cliente

    top_deudores = sorted(
        deuda_rows, key=lambda x: x['total'], reverse=True
    )[:10]
    deuda_ids = {row['cliente'].id for row in deuda_rows}

    good_payers = (
        Pago.objects.values('cliente')
        .annotate(total=Sum('monto'))
        .order_by('-total')
    )
    good_payer_ids = [item['cliente'] for item in good_payers]
    good_payer_lookup = {
        item['cliente']: item['total'] for item in good_payers
    }
    good_clients = []
    for c in Cliente.objects.filter(id__in=good_payer_ids):
        cliente_id = cast(Any, c).id
        if cliente_id in deuda_ids:
            continue
        good_clients.append({
            'cliente': c,
            'total': good_payer_lookup.get(cliente_id, 0)
        })
    good_clients = sorted(
        good_clients, key=lambda x: x['total'], reverse=True
    )[:10]

    return render(
        request,
        'billing_app/reportes/index.html',
        {
            'ingresos_total': ingresos_total,
            'egresos_total': egresos_total,
            'total_clientes': total_clientes,
            'total_planes': total_planes,
            'total_servicios': total_servicios,
            'total_nodos': total_nodos,
            'instalaciones': instalaciones,
            'cortes': cortes,
            'averias': averias,
            'ots_completadas': ots_completadas,
            'clientes_por_zona': clientes_por_zona,
            'zonas_por_sector': zonas_por_sector,
            'deuda_total': deuda_total,
            'top_deudores': top_deudores,
            'good_clients': good_clients
        }
    )


@login_required(login_url='admin:login')
def reportes_ingresos_egresos(request):
    pagos = (
        Pago.objects.select_related('cliente')
        .prefetch_related(
            'items__plan_asociado__plan__servicio',
            'items__ot_asociada__concepto'
        )
        .order_by('-fecha')
    )
    egresos = (
        Egreso.objects.select_related('concepto')
        .order_by('-fecha')
    )

    ingresos_rows = []
    for pago in pagos:
        items = list(cast(Any, pago).items.all())
        if not items:
            ingresos_rows.append([
                pago.fecha, str(pago.cliente), pago.cliente.dni,
                pago.tipo_comprobante, pago.serie_numero,
                pago.monto, pago.detalles, '', '', '', '', ''
            ])
            continue
        for item in items:
            plan = item.plan_asociado.plan if item.plan_asociado else None
            servicio = plan.servicio if plan and plan.servicio else None
            ot = item.ot_asociada
            plan_label = ''
            if plan:
                plan_label = plan.nombre
                if servicio:
                    plan_label = f"{servicio.nombre} - {plan.nombre}"
            ingresos_rows.append([
                pago.fecha, str(pago.cliente), pago.cliente.dni,
                pago.tipo_comprobante, pago.serie_numero,
                pago.monto, pago.detalles,
                item.descripcion, item.monto_parcial,
                item.periodo_mes,
                plan_label,
                f"OT-{ot.id} {ot.concepto.nombre}" if ot else ''
            ])

    egresos_rows = [
        [
            egreso.fecha,
            egreso.concepto.nombre,
            egreso.monto,
            egreso.tipo_comprobante,
            egreso.numero_comprobante,
            egreso.responsable,
            egreso.observaciones,
            egreso.created_at
        ]
        for egreso in egresos
    ]

    return _excel_response(
        'reporte_ingresos_egresos.xlsx',
        [
            {
                'title': 'Ingresos',
                'headers': [
                    'Fecha', 'Cliente', 'DNI', 'Tipo comprobante',
                    'Serie/Numero', 'Monto pago', 'Detalle pago',
                    'Detalle item', 'Monto item', 'Periodo mes',
                    'Plan', 'OT'
                ],
                'rows': ingresos_rows
            },
            {
                'title': 'Egresos',
                'headers': [
                    'Fecha', 'Concepto', 'Monto', 'Tipo comprobante',
                    'Numero comprobante', 'Responsable',
                    'Observaciones', 'Creado'
                ],
                'rows': egresos_rows
            }
        ]
    )


@login_required(login_url='admin:login')
def reportes_clientes(request):
    clientes = (
        Cliente.objects.select_related(
            'via__sector__distrito', 'mikrotik_vinculado'
        )
        .order_by('apellidos', 'nombres')
    )

    rows = []
    for cliente in clientes:
        via = cliente.via
        sector = via.sector if via else None
        distrito = sector.distrito if sector else None
        via_disp = ''
        if via:
            via_disp = f"{cast(Any, via).get_tipo_display()} {via.nombre}"
        rows.append([
            cliente.apellidos,
            cliente.nombres,
            cliente.dni,
            cliente.celular,
            distrito.nombre if distrito else '',
            sector.nombre if sector else '',
            via_disp,
            cliente.referencia or '',
            cliente.estado_activo,
            cliente.tipo_conexion,
            cliente.usuario_pppoe or '',
            cliente.ip_asignada or '',
            (
                cliente.mikrotik_vinculado.nombre
                if cliente.mikrotik_vinculado else ''
            ),
            cliente.fecha_registro
        ])

    return _excel_response(
        'reporte_clientes.xlsx',
        [
            {
                'title': 'Clientes',
                'headers': [
                    'Apellidos', 'Nombres', 'DNI', 'Celular',
                    'Distrito', 'Sector', 'Via', 'Referencia',
                    'Activo', 'Tipo conexion', 'Usuario PPPoE',
                    'IP asignada', 'Mikrotik', 'Fecha registro'
                ],
                'rows': rows
            }
        ]
    )


@login_required(login_url='admin:login')
def reportes_ots(request):
    ots = (
        OrdenTecnica.objects.select_related(
            'cliente', 'concepto', 'tecnico_asignado',
            'servicio_afectado', 'plan_asociado__plan__servicio'
        )
        .order_by('-fecha_creacion')
    )

    rows = []
    for ot in ots:
        plan = ot.plan_asociado.plan if ot.plan_asociado else None
        servicio = plan.servicio if plan and plan.servicio else None
        ot_id = cast(Any, ot).id
        concepto = cast(Any, ot.concepto)
        rows.append([
            ot_id,
            str(ot.cliente),
            ot.cliente.dni,
            concepto.nombre,
            concepto.get_categoria_display(),
            (
                servicio.nombre if servicio
                else (
                    ot.servicio_afectado.nombre
                    if ot.servicio_afectado else ''
                )
            ),
            plan.nombre if plan else '',
            ot.monto,
            ot.observaciones or '',
            ot.completada,
            ot.pagada,
            ot.exonerada,
            ot.fecha_creacion,
            ot.fecha_finalizacion,
            ot.tecnico_asignado.nombre if ot.tecnico_asignado else ''
        ])

    return _excel_response(
        'reporte_ots.xlsx',
        [
            {
                'title': 'Ordenes Tecnicas',
                'headers': [
                    'OT', 'Cliente', 'DNI', 'Concepto', 'Categoria',
                    'Servicio', 'Plan', 'Monto', 'Observaciones',
                    'Completada', 'Pagada', 'Exonerada',
                    'Fecha creacion', 'Fecha finalizacion', 'Tecnico'
                ],
                'rows': rows
            }
        ]
    )


@login_required(login_url='admin:login')
def reportes_clientes_planes(request):
    planes = (
        ClientePlan.objects.select_related(
            'cliente', 'plan__servicio'
        )
        .order_by('cliente__apellidos', 'cliente__nombres')
    )

    rows = []
    for cliente_plan in planes:
        plan = cliente_plan.plan
        servicio = plan.servicio if plan and plan.servicio else None
        cliente = cliente_plan.cliente
        rows.append([
            str(cliente),
            cliente.dni,
            plan.nombre if plan else '',
            servicio.nombre if servicio else '',
            plan.precio if plan else '',
            cliente_plan.fecha_inicio,
            cliente_plan.fecha_cobranza,
            cliente.estado_activo,
            cliente.tipo_conexion,
            cliente.usuario_pppoe or '',
            cliente.ip_asignada or '',
            (
                cliente.mikrotik_vinculado.nombre
                if cliente.mikrotik_vinculado else ''
            )
        ])

    return _excel_response(
        'reporte_clientes_con_planes.xlsx',
        [
            {
                'title': 'Clientes con planes',
                'headers': [
                    'Cliente', 'DNI', 'Plan', 'Servicio', 'Precio',
                    'Fecha inicio', 'Dia cobranza', 'Activo',
                    'Tipo conexion', 'Usuario PPPoE', 'IP asignada',
                    'Mikrotik'
                ],
                'rows': rows
            }
        ]
    )


@login_required(login_url='admin:login')
def reportes_deuda(request):
    rows = []
    for cliente in Cliente.objects.all().order_by('apellidos', 'nombres'):
        cliente_id = cast(Any, cliente).id
        try:
            deuda = calcular_meses_deuda(cliente)
        except Exception:
            logger.exception(
                'Error calculando deuda para cliente %s',
                cliente_id
            )
            deuda = []
        total_cliente = sum(
            (_to_decimal(item.get('saldo')) for item in deuda),
            Decimal('0')
        )
        if total_cliente <= 0:
            continue
        rows.append([
            str(cliente),
            cliente.dni,
            cliente.celular,
            cliente.estado_activo,
            total_cliente
        ])

    return _excel_response(
        'reporte_deuda_pendiente.xlsx',
        [
            {
                'title': 'Deuda pendiente',
                'headers': [
                    'Cliente', 'DNI', 'Celular', 'Activo', 'Deuda total'
                ],
                'rows': rows
            }
        ]
    )


@login_required(login_url='admin:login')
def cliente_lista(request):
    """Lista de clientes con búsqueda y filtros"""
    search_query = request.GET.get('q', '')
    clientes = (
        Cliente.objects.all()
        .prefetch_related('planes', 'ordenes_tecnicas')
    )
    if search_query:
        clientes = clientes.filter(
            Q(apellidos__icontains=search_query) |
            Q(nombres__icontains=search_query) |
            Q(dni__icontains=search_query)
        )
    return render(
        request,
        'billing_app/index.html',
        {'clientes': clientes, 'search_query': search_query}
    )


def cliente_generar_form(request, form, title, edit_mode=False):
    """Renderiza el formulario de cliente"""
    return render(
        request,
        'billing_app/cliente_form.html',
        {'form': form, 'edit_mode': edit_mode, 'titulo': title}
    )


@login_required(login_url='admin:login')
def cliente_crear(request):
    """Crea nuevo cliente"""
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = ClienteForm()
    return cliente_generar_form(request, form, 'Nuevo Cliente')


@login_required(login_url='admin:login')
def cliente_editar(request, pk):
    """Edita cliente existente"""
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect('cliente-detalle', pk=cliente.pk)
    else:
        form = ClienteForm(instance=cliente)
    return cliente_generar_form(
        request, form, 'Editar Cliente', edit_mode=True
    )


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def cliente_eliminar(request, pk):
    if not can_delete_cliente(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.delete()
    return redirect('index')


def _construir_historial_cliente(cliente):
    """Lista de eventos (cortes, planes, instalación, pagos) ordenados
    por fecha desc."""
    eventos = []
    for mov in (
        cliente.movimientos_historial.all()
        .order_by('-fecha')[:200]
    ):
        eventos.append({
            'fecha': mov.fecha.date(),
            'fecha_dt': mov.fecha,
            'tipo': 'movimiento',
            'titulo': mov.tipo,
            'detalle': mov.detalle,
            'icono': mov.icono,
            'clase': mov.clase
        })
    for cp in _cliente_planes_qs(cliente).all().select_related('plan'):
        fd = _to_aware_datetime(cp.fecha_inicio)
        eventos.append({
            'fecha': cp.fecha_inicio,
            'fecha_dt': fd,
            'tipo': 'plan',
            'titulo': 'Plan asignado / Instalación',
            'detalle': (
                f"{cp.plan.nombre} (S/ {cp.plan.precio}), "
                f"cobranza día {cp.fecha_cobranza}"
            ),
            'icono': 'fa-wifi',
            'clase': 'success'
        })
    for p in cliente.pagos.all().order_by('-fecha')[:200]:
        fecha_dt = _to_aware_datetime(p.fecha)
        eventos.append({
            'fecha': p.fecha.date() if hasattr(p.fecha, 'date') else p.fecha,
            'fecha_dt': fecha_dt,
            'tipo': 'pago',
            'titulo': 'Pago',
            'detalle': f"{p.tipo_comprobante} {p.serie_numero} - S/ {p.monto}",
            'icono': 'fa-money-bill-wave',
            'clase': 'primary'
        })
    for ot in (
        cliente.ordenes_tecnicas.all()
        .select_related('concepto')
        .order_by('-fecha_creacion')
    ):
        f = ot.fecha_finalizacion or ot.fecha_creacion
        fecha_dt = _to_aware_datetime(f)
        nombre = (ot.concepto.nombre or '').lower()
        if 'corte' in nombre or ot.concepto.categoria == 'CORTES':
            tipo_label = 'Corte'
            icono = 'fa-ban'
            clase = 'danger'
        elif 'reconexion' in nombre or 'reconexi?n' in nombre:
            tipo_label = 'Reconexión'
            icono = 'fa-plug'
            clase = 'success'
        else:
            tipo_label = 'Orden técnica'
            icono = 'fa-tools'
            clase = 'warning'
        status_label = (
            ' (completada)' if ot.completada else ' (pendiente)'
        )
        eventos.append({
            'fecha': f.date() if hasattr(f, 'date') else f,
            'fecha_dt': fecha_dt,
            'tipo': 'ot',
            'titulo': tipo_label,
            'detalle': (
                f"{ot.concepto.nombre} - S/ {ot.monto}{status_label}"
            ),
            'icono': icono,
            'clase': clase
        })
    eventos.sort(
        key=lambda x: x['fecha_dt'] or timezone.now(),
        reverse=True
    )
    return eventos[:150]


@login_required(login_url='admin:login')
def cliente_detalle(request, pk):
    """Detalle completo del cliente con deuda, OTs e historial de
    movimientos."""
    cliente = get_object_or_404(Cliente, pk=pk)
    try:
        deuda = calcular_meses_deuda(cliente)
    except Exception:
        deuda = []
        logger.exception("Error calculando deuda para cliente %s", cliente.pk)

    total_deuda = sum(
        (_to_decimal(item.get('saldo')) for item in deuda),
        Decimal('0')
    )
    ots = OrdenTecnica.objects.filter(
        cliente=cliente
    ).select_related('concepto', 'tecnico_asignado').order_by(
        '-fecha_creacion'
    )
    historial = _construir_historial_cliente(cliente)
    mikrotik_id = None
    if cliente.mikrotik_vinculado:
        mikrotik_id = cast(Any, cliente.mikrotik_vinculado).id
    connected_ids = set(
        (request.session.get('mikrotik_connections') or {}).keys()
    )
    mikrotik_connected = (
        mikrotik_id is not None
        and str(mikrotik_id) in connected_ids
    )

    return render(request, 'billing_app/cliente_detalle.html', {
        'cliente': cliente,
        'ots': ots,
        'deuda': deuda,
        'total_deuda': total_deuda,
        'historial': historial,
        'mikrotik_id': mikrotik_id,
        'mikrotik_connected': mikrotik_connected
    })


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def cliente_status_toggle(request, pk):
    """Alterna estado activo/inactivo del cliente"""
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente.estado_activo = not cliente.estado_activo
    cliente.save()
    registrar_movimiento(
        cliente,
        'Estado actualizado',
        'Cliente activado' if cliente.estado_activo else 'Cliente desactivado',
        icono='fa-user-check' if cliente.estado_activo else 'fa-user-slash',
        clase='success' if cliente.estado_activo else 'danger'
    )
    return redirect('cliente-detalle', pk=cliente.pk)


@login_required(login_url='admin:login')
def ajax_load_sectores(request):
    """AJAX: Carga sectores según distrito seleccionado"""
    distrito_id = request.GET.get('distrito')
    if not distrito_id:
        return JsonResponse([], safe=False)
    sectores = Sector.objects.filter(
        distrito_id=distrito_id
    ).values('id', 'nombre')
    return JsonResponse(list(sectores), safe=False)


@login_required(login_url='admin:login')
def ajax_load_vias(request):
    """AJAX: Carga vías según sector seleccionado"""
    sector_id = request.GET.get('sector')
    if not sector_id:
        return JsonResponse([], safe=False)
    vias = Via.objects.filter(
        sector_id=sector_id
    ).values('id', 'nombre', 'tipo')
    vias_list = []
    for v in vias:
        tipo_disp = dict(Via.TIPO_CHOICES).get(v['tipo'], '')
        vias_list.append({
            'id': v['id'],
            'nombre': f"{tipo_disp} {v['nombre']}"
        })
    return JsonResponse(vias_list, safe=False)


@login_required(login_url='admin:login')
def cliente_agregar_plan(request, pk):
    """Agrega un nuevo plan a un cliente; muestra catálogo de planes."""
    cliente = get_object_or_404(Cliente, pk=pk)
    planes = Plan.objects.select_related('servicio').order_by(
        'servicio__nombre', 'nombre'
    )
    if request.method == 'POST':
        form = ClientePlanForm(request.POST)
        if form.is_valid():
            cliente_plan = form.save(commit=False)
            cliente_plan.cliente = cliente
            cliente_plan.save()
            registrar_movimiento(
                cliente,
                'Plan creado',
                f"Se asignó el plan {cliente_plan.plan.nombre}",
                icono='fa-wifi',
                clase='success'
            )
            return redirect('cliente-detalle', pk=cliente.pk)
    else:
        form = ClientePlanForm()
    return render(
        request,
        'billing_app/ajustes/partial_form_plan.html',
        {
            'form': form,
            'cliente': cliente,
            'titulo': 'Añadir Plan',
            'planes': planes
        }
    )


@login_required(login_url='admin:login')
def cliente_editar_plan(request, pk, plan_id):
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente_plan = get_object_or_404(
        ClientePlan, pk=plan_id, cliente=cliente
    )
    planes = Plan.objects.select_related('servicio').order_by(
        'servicio__nombre', 'nombre'
    )
    if request.method == 'POST':
        form = ClientePlanForm(request.POST, instance=cliente_plan)
        if form.is_valid():
            actualizado = form.save()
            registrar_movimiento(
                cliente,
                'Plan actualizado',
                f"Plan actualizado: {actualizado.plan.nombre}",
                icono='fa-pen',
                clase='info'
            )
            return redirect('cliente-detalle', pk=cliente.pk)
    else:
        form = ClientePlanForm(instance=cliente_plan)
    return render(
        request,
        'billing_app/ajustes/partial_form_plan.html',
        {
            'form': form,
            'cliente': cliente,
            'titulo': 'Editar Plan',
            'planes': planes
        }
    )


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def cliente_eliminar_plan(request, pk, plan_id):
    cliente = get_object_or_404(Cliente, pk=pk)
    cliente_plan = get_object_or_404(
        ClientePlan, pk=plan_id, cliente=cliente
    )
    plan_nombre = cliente_plan.plan.nombre
    cliente_plan.delete()
    registrar_movimiento(
        cliente,
        'Plan eliminado',
        f"Se eliminó el plan {plan_nombre}",
        icono='fa-trash-alt',
        clase='danger'
    )
    return redirect('cliente-detalle', pk=cliente.pk)


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def cliente_deuda_eliminar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    tipo = (request.POST.get('tipo') or '').upper()
    plan_id = request.POST.get('plan_id')
    mes_iso = request.POST.get('mes_iso')

    if tipo == 'OT' and plan_id:
        ot = get_object_or_404(OrdenTecnica, pk=plan_id, cliente=cliente)
        registrar_movimiento(
            cliente,
            'Deuda OT eliminada',
            f"OT eliminada desde estado de cuenta: {ot.concepto.nombre}",
            icono='fa-trash-alt',
            clase='danger'
        )
        ot.delete()
        return redirect('cliente-detalle', pk=cliente.pk)

    if tipo == 'PLAN' and plan_id and mes_iso:
        try:
            periodo = datetime.strptime(mes_iso[:10], '%Y-%m-%d').date()
        except ValueError:
            periodo = None
        plan = get_object_or_404(ClientePlan, pk=plan_id, cliente=cliente)
        if periodo:
            DeudaExcluida.objects.get_or_create(
                cliente=cliente,
                plan_asociado=plan,
                periodo_mes=periodo,
                defaults={'motivo': 'Eliminado desde estado de cuenta'}
            )
            registrar_movimiento(
                cliente,
                'Deuda de plan eliminada',
                f"Se excluyó {plan.plan.nombre} ({periodo.strftime('%m/%Y')})",
                icono='fa-trash-alt',
                clase='danger'
            )
    return redirect('cliente-detalle', pk=cliente.pk)


@login_required(login_url='admin:login')
def api_generar_id_pppoe(request):
    """Genera ID PPPoE único para cliente"""
    max_attempts = 100
    for _ in range(max_attempts):
        nuevo_id = str(random.randint(1000, 9999))
        if not Cliente.objects.filter(usuario_pppoe=nuevo_id).exists():
            return JsonResponse({'id': nuevo_id})
    return JsonResponse({'error': 'No se pudo generar ID único'}, status=500)


@login_required(login_url='admin:login')
def cliente_crear_ot(request, pk):
    """Crea orden tecnica para un cliente (panel lateral)."""
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        post_data = request.POST
        tipo_trabajo = (post_data.get('tipo_trabajo') or '').upper()
        concepto_id = post_data.get('concepto')
        if not tipo_trabajo and concepto_id:
            concepto_sel = OrdenTecnicaConcepto.objects.filter(
                id=concepto_id
            ).first()
            if concepto_sel:
                mutable = post_data.copy()
                mutable['tipo_trabajo'] = concepto_sel.categoria
                post_data = mutable
                tipo_trabajo = concepto_sel.categoria
        if tipo_trabajo in ('AVERIAS', 'CORTES') and not concepto_id:
            concepto_qs = OrdenTecnicaConcepto.objects.filter(
                categoria=tipo_trabajo
            ).order_by('nombre')
            concepto_first = concepto_qs.first()
            if concepto_first:
                mutable = post_data.copy()
                mutable['concepto'] = str(cast(Any, concepto_first).id)
                post_data = mutable
        if tipo_trabajo in ('AVERIAS', 'CORTES'):
            planes_sel = post_data.getlist('planes_seleccionados')
            if not planes_sel:
                planes_qs = _cliente_planes_qs(cliente)
                if planes_qs.count() == 1:
                    mutable = post_data.copy()
                    mutable.setlist(
                        'planes_seleccionados',
                        [str(planes_qs.first().id)]
                    )
                    post_data = mutable
        form = OrdenTecnicaForm(post_data)
        if 'plan_asociado' in form.fields:
            field = cast(ModelChoiceField, form.fields['plan_asociado'])
            field.queryset = _cliente_planes_qs(cliente).select_related(
                'plan', 'plan__servicio'
            ).all()
        if 'planes_seleccionados' in form.fields:
            field = cast(ModelMultipleChoiceField,
                         form.fields['planes_seleccionados'])
            field.queryset = _cliente_planes_qs(cliente).select_related(
                'plan', 'plan__servicio'
            ).all()
        if 'plan_catalogo' in form.fields:
            field = cast(
                ModelMultipleChoiceField,
                form.fields['plan_catalogo']
            )
            field.queryset = Plan.objects.select_related('servicio').all()
        if 'servicio_afectado' in form.fields:
            field = cast(ModelChoiceField, form.fields['servicio_afectado'])
            field.queryset = Servicio.objects.all().order_by('nombre')
        if 'servicios_seleccionados' in form.fields:
            field = cast(ModelMultipleChoiceField,
                         form.fields['servicios_seleccionados'])
            field.queryset = _cliente_servicios_qs(cliente)
        if form.is_valid():
            concepto = form.cleaned_data['concepto']
            tipo_trabajo = form.cleaned_data.get('tipo_trabajo')
            tecnico = form.cleaned_data.get('tecnico_asignado')
            observaciones = form.cleaned_data.get('observaciones')
            monto = form.cleaned_data.get('monto')
            creadas = []

            if tipo_trabajo == 'INSTALACION':
                planes = form.cleaned_data.get('plan_catalogo') or []
                for plan in planes:
                    obs_plan = f"Instalación de plan {plan.nombre}"
                    if observaciones:
                        obs_plan = f"{obs_plan} | {observaciones}"
                    cliente_plan, creado = ClientePlan.objects.get_or_create(
                        cliente=cliente,
                        plan=plan,
                        defaults={
                            'fecha_inicio': timezone.now().date(),
                            'fecha_cobranza': timezone.now().day,
                            'activo': True
                        }
                    )
                    if not cliente_plan.activo:
                        cliente_plan.activo = True
                        cliente_plan.save(update_fields=['activo'])
                    if creado:
                        registrar_movimiento(
                            cliente,
                            'Plan creado',
                            f"Se asignó el plan {plan.nombre}",
                            icono='fa-wifi',
                            clase='success'
                        )
                    ot = OrdenTecnica.objects.create(
                        cliente=cliente,
                        concepto=concepto,
                        plan_asociado=None,
                        servicio_afectado=plan.servicio if plan else None,
                        monto=monto,
                        observaciones=obs_plan,
                        tecnico_asignado=tecnico
                    )
                    creadas.append(ot)
            elif tipo_trabajo in ('AVERIAS', 'CORTES'):
                planes = list(
                    form.cleaned_data.get('planes_seleccionados') or []
                )
                planes_catalogo = list(
                    form.cleaned_data.get('plan_catalogo') or []
                )
                if not planes and not planes_catalogo:
                    form.add_error(
                        None,
                        'Seleccione planes activos para la OT.'
                    )
                if not planes and planes_catalogo:
                    cliente_planes = (
                        _cliente_planes_qs(cliente)
                        .select_related('plan', 'plan__servicio')
                        .all()
                    )
                    cliente_plan_lookup = {
                        cp.plan_id: cp for cp in cliente_planes
                    }
                    for plan in planes_catalogo:
                        plan_cliente = cliente_plan_lookup.get(plan.id)
                        if (
                            plan_cliente
                            and getattr(plan_cliente, 'activo', True)
                        ):
                            planes.append(plan_cliente)
                            continue
                        obs_plan = f"Plan: {plan.nombre}"
                        if observaciones:
                            obs_plan = f"{obs_plan} | {observaciones}"
                        ot = OrdenTecnica.objects.create(
                            cliente=cliente,
                            concepto=concepto,
                            plan_asociado=None,
                            servicio_afectado=(
                                plan.servicio if plan else None
                            ),
                            monto=monto,
                            observaciones=obs_plan,
                            tecnico_asignado=tecnico
                        )
                        creadas.append(ot)
                for plan in planes:
                    plan_name = (
                        plan.plan.nombre if plan and plan.plan else 'Plan'
                    )
                    obs_plan = f"Plan: {plan_name}"
                    if observaciones:
                        obs_plan = f"{obs_plan} | {observaciones}"
                    ot = OrdenTecnica.objects.create(
                        cliente=cliente,
                        concepto=concepto,
                        plan_asociado=plan,
                        servicio_afectado=(
                            plan.plan.servicio if plan else None
                        ),
                        monto=monto,
                        observaciones=obs_plan,
                        tecnico_asignado=tecnico
                    )
                    creadas.append(ot)

            if form.errors:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return render(
                        request,
                        'billing_app/cliente_ot_form.html',
                        {
                            'form': form,
                            'cliente': cliente,
                            'conceptos': (
                                OrdenTecnicaConcepto.objects
                                .all()
                                .order_by('categoria', 'nombre')
                            ),
                            'cat_display': dict(
                                OrdenTecnicaConcepto.CAT_CHOICES
                            ),
                            'cliente_planes': (
                                _cliente_planes_qs(cliente)
                                .select_related('plan', 'plan__servicio')
                                .all()
                            ),
                            'cliente_servicios': _cliente_servicios_qs(
                                cliente
                            ),
                            'planes_catalogo': Plan.objects.select_related(
                                'servicio'
                            ).all()
                        },
                        status=400
                    )
                return redirect('cliente-detalle', pk=cliente.pk)

            for ot in creadas:
                registrar_movimiento(
                    cliente,
                    'OT creada',
                    f"{ot.concepto.nombre} - S/ {ot.monto}",
                    icono='fa-tools',
                    clase='warning'
                )
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'ok': True})
            return redirect('cliente-detalle', pk=cliente.pk)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return render(
                request,
                'billing_app/cliente_ot_form.html',
                {
                    'form': form,
                    'cliente': cliente,
                    'conceptos': OrdenTecnicaConcepto.objects.all().order_by(
                        'categoria', 'nombre'
                    ),
                    'cat_display': dict(OrdenTecnicaConcepto.CAT_CHOICES),
                    'cliente_planes': (
                        _cliente_planes_qs(cliente)
                        .select_related('plan', 'plan__servicio')
                        .all()
                    ),
                    'cliente_servicios': _cliente_servicios_qs(cliente),
                    'planes_catalogo': Plan.objects.select_related(
                        'servicio'
                    ).all()
                },
                status=400
            )
    else:
        form = OrdenTecnicaForm()
        if 'plan_asociado' in form.fields:
            field = cast(ModelChoiceField, form.fields['plan_asociado'])
            field.queryset = _cliente_planes_qs(cliente).select_related(
                'plan', 'plan__servicio'
            ).all()
        if 'planes_seleccionados' in form.fields:
            field = cast(ModelMultipleChoiceField,
                         form.fields['planes_seleccionados'])
            field.queryset = _cliente_planes_qs(cliente).select_related(
                'plan', 'plan__servicio'
            ).all()
        if 'plan_catalogo' in form.fields:
            field = cast(
                ModelMultipleChoiceField,
                form.fields['plan_catalogo']
            )
            field.queryset = Plan.objects.select_related('servicio').all()
        if 'servicio_afectado' in form.fields:
            field = cast(ModelChoiceField, form.fields['servicio_afectado'])
            field.queryset = Servicio.objects.all().order_by('nombre')
        if 'servicios_seleccionados' in form.fields:
            field = cast(ModelMultipleChoiceField,
                         form.fields['servicios_seleccionados'])
            field.queryset = _cliente_servicios_qs(cliente)
    return render(
        request,
        'billing_app/cliente_ot_form.html',
        {
            'form': form,
            'cliente': cliente,
            'conceptos': OrdenTecnicaConcepto.objects.all().order_by(
                'categoria', 'nombre'
            ),
            'cat_display': dict(OrdenTecnicaConcepto.CAT_CHOICES),
            'cliente_planes': _cliente_planes_qs(cliente).select_related(
                'plan', 'plan__servicio'
            ).all(),
            'cliente_servicios': _cliente_servicios_qs(cliente),
            'planes_catalogo': Plan.objects.select_related('servicio').all()
        }
    )


@login_required(login_url='admin:login')
def cliente_ot_crear(request, pk):
    """Crea orden técnica para cliente (GET: datos, POST: crear OT)"""
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        try:
            tipo_trabajo = (request.POST.get('tipo_trabajo') or '').upper()
            concepto_id = request.POST.get('concepto')
            observaciones = (
                request.POST.get('observaciones', '').strip()
            )
            fecha_asistencia_raw = (
                request.POST.get('fecha_asistencia') or ''
            ).strip()
            fecha_asistencia = None
            if fecha_asistencia_raw:
                fecha_asistencia = parse_date(fecha_asistencia_raw)
            if fecha_asistencia_raw and not fecha_asistencia:
                msg = {'ok': False, 'error': 'Fecha inválida'}
                return JsonResponse(msg, status=400)

            # Validar tipo y concepto
            cat_choices = dict(OrdenTecnicaConcepto.CAT_CHOICES)
            if not tipo_trabajo or tipo_trabajo not in cat_choices:
                msg = {'ok': False, 'error': 'Tipo inválido'}
                return JsonResponse(msg, status=400)
            if not concepto_id:
                msg = {'ok': False, 'error': 'Selecciona concepto'}
                return JsonResponse(msg, status=400)

            concepto = OrdenTecnicaConcepto.objects.filter(
                pk=concepto_id
            ).first()
            if not concepto:
                msg = {'ok': False, 'error': 'Concepto no existe'}
                return JsonResponse(msg, status=400)

            creadas = []

            if tipo_trabajo == 'INSTALACION':
                # Obtener planes seleccionados del catálogo
                planes_ids = request.POST.getlist('planes_catalogo')
                if not planes_ids:
                    msg = {'ok': False, 'error': 'Selecciona un plan'}
                    return JsonResponse(msg, status=400)

                planes = Plan.objects.filter(id__in=planes_ids)
                for plan in planes:
                    # Crear/obtener ClientePlan
                    cp, _ = ClientePlan.objects.get_or_create(
                        cliente=cliente,
                        plan=plan,
                        defaults={
                            'fecha_inicio': timezone.now().date(),
                            'fecha_cobranza': timezone.now().day,
                            'activo': True
                        }
                    )
                    if not cp.activo:
                        cp.activo = True
                        cp.save(update_fields=['activo'])

                    obs = f"Instalación de {plan.nombre}"
                    if observaciones:
                        obs += f" | {observaciones}"

                    # Crear OT
                    ot = OrdenTecnica.objects.create(
                        cliente=cliente,
                        concepto=concepto,
                        plan_asociado=cp,
                        monto=concepto.precio_sugerido or Decimal('0'),
                        observaciones=obs,
                        fecha_asistencia=fecha_asistencia
                    )
                    creadas.append(ot)

            elif tipo_trabajo in ('AVERIAS', 'CORTES', 'RECONEXION'):
                # Obtener planes seleccionados del cliente
                planes_ids = request.POST.getlist('planes_cliente')
                if not planes_ids:
                    msg = {'ok': False, 'error': 'Selecciona un plan'}
                    return JsonResponse(msg, status=400)

                planes = ClientePlan.objects.filter(
                    id__in=planes_ids,
                    cliente=cliente
                )
                if tipo_trabajo in ('AVERIAS', 'CORTES'):
                    planes = planes.filter(activo=True)
                elif tipo_trabajo == 'RECONEXION':
                    planes = planes.filter(activo=False)
                if not planes.exists():
                    msg = {
                        'ok': False,
                        'error': 'Selecciona planes con estado valido'
                    }
                    return JsonResponse(msg, status=400)

                for cp in planes:
                    obs = f"Plan: {cp.plan.nombre}"
                    if observaciones:
                        obs += f" | {observaciones}"

                    # Crear OT
                    ot = OrdenTecnica.objects.create(
                        cliente=cliente,
                        concepto=concepto,
                        plan_asociado=cp,
                        monto=concepto.precio_sugerido or Decimal('0'),
                        observaciones=obs,
                        fecha_asistencia=fecha_asistencia
                    )
                    creadas.append(ot)

            # Registrar movimientos
            for ot in creadas:
                registrar_movimiento(
                    cliente,
                    'OT creada',
                    f"{ot.concepto.nombre}",
                    icono='fa-tools',
                    clase='warning'
                )

            return JsonResponse({
                'ok': True,
                'count': len(creadas),
                'message': f'{len(creadas)} OT(s) creada(s) correctamente'
            })
        except Exception as e:
            return JsonResponse({
                'ok': False,
                'error': f'Error en servidor: {str(e)}'
            }, status=500)

    # GET: Retornar datos para el formulario
    tipo_param = (request.GET.get('tipo') or '').upper()

    response = {
        'ok': True,
        'cliente_id': cliente.pk,
        'tipo': tipo_param
    }

    # Conceptos filtrados por tipo
    if tipo_param:
        conceptos = OrdenTecnicaConcepto.objects.filter(
            categoria=tipo_param
        ).order_by('nombre')
    else:
        conceptos = OrdenTecnicaConcepto.objects.all().order_by(
            'categoria', 'nombre'
        )

    response['conceptos'] = [
        {
            'id': c.pk,
            'nombre': c.nombre,
            'precio': float(c.precio_sugerido or 0)
        }
        for c in conceptos
    ]

    # Planes según tipo
    if tipo_param == 'INSTALACION':
        planes = Plan.objects.select_related(
            'servicio'
        ).all().order_by('nombre')
        response['planes'] = [
            {
                'id': p.pk,
                'nombre': p.nombre,
                'servicio': p.servicio.nombre if p.servicio else '',
                'precio': float(p.precio)
            }
            for p in planes
        ]
    elif tipo_param in ('AVERIAS', 'CORTES'):
        planes = ClientePlan.objects.filter(
            cliente=cliente,
            activo=True
        ).select_related('plan', 'plan__servicio').order_by('plan__nombre')
        response['planes'] = [
            {
                'id': p.pk,
                'nombre': p.plan.nombre,
                'servicio': p.plan.servicio.nombre if p.plan.servicio else '',
                'tipo': 'cliente_plan',
                'activo': p.activo
            }
            for p in planes
        ]
    elif tipo_param == 'RECONEXION':
        # Mostrar solo planes inactivos del cliente
        planes = ClientePlan.objects.filter(
            cliente=cliente,
            activo=False
        ).select_related('plan', 'plan__servicio').order_by('plan__nombre')
        response['planes'] = [
            {
                'id': p.pk,
                'nombre': p.plan.nombre,
                'servicio': p.plan.servicio.nombre if p.plan.servicio else '',
                'tipo': 'cliente_plan',
                'activo': p.activo
            }
            for p in planes
        ]

    return JsonResponse(response)


# --- Tecnicos ---

@login_required(login_url='admin:login')
def tecnico_lista(request):
    """Lista de técnicos con órdenes técnicas pendientes y realizadas"""
    tecnicos = Tecnico.objects.filter(
        activo=True
    ).prefetch_related('ordenes')
    ots_pendientes = OrdenTecnica.objects.filter(
        completada=False
    ).select_related(
        'cliente', 'concepto', 'tecnico_asignado'
    ).order_by('-fecha_creacion')
    ots_realizadas = OrdenTecnica.objects.filter(
        completada=True
    ).select_related(
        'cliente', 'concepto', 'tecnico_asignado'
    ).order_by('-fecha_finalizacion', '-fecha_creacion')[:100]
    return render(request, 'billing_app/tecnico_lista.html', {
        'tecnicos': tecnicos,
        'ots_pendientes': ots_pendientes,
        'ots_realizadas': ots_realizadas
    })


@login_required(login_url='admin:login')
def tecnico_crear(request):
    """Crea nuevo técnico"""
    return generic_side_panel_form(
        request, TecnicoForm, 'Nuevo Técnico', 'tecnico-lista'
    )


@login_required(login_url='admin:login')
def tecnico_editar(request, pk):
    """Edita técnico existente"""
    obj = get_object_or_404(Tecnico, pk=pk)
    return generic_side_panel_form(
        request, TecnicoForm, 'Editar Técnico', 'tecnico-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def tecnico_eliminar(request, pk):
    get_object_or_404(Tecnico, pk=pk).delete()
    return redirect('tecnico-lista')


# --- OT Gestion ---

def _aplicar_estado_cliente_por_ot(ot):
    """
    Aplica cambios de estado según el tipo de OT:
    - INSTALACION: Activa el plan y el cliente
    - CORTES: Desactiva el plan, cliente inactivo solo si TODOS
      los planes están inactivos
    - RECONEXION: Reactiva el plan y el cliente
    """
    nombre_concepto = (ot.concepto.nombre or '').strip().lower()
    categoria = (ot.concepto.categoria or '').upper()

    if categoria == 'INSTALACION':
        # Activar el plan asociado
        if ot.plan_asociado:
            ot.plan_asociado.activo = True
            ot.plan_asociado.save(update_fields=['activo'])
        # Activar cliente
        ot.cliente.estado_activo = True
        ot.cliente.save(update_fields=['estado_activo'])

    elif 'corte' in nombre_concepto or categoria == 'CORTES':
        # Desactivar el plan asociado
        if ot.plan_asociado:
            ot.plan_asociado.activo = False
            ot.plan_asociado.save(update_fields=['activo'])

        # Cliente inactivo solo si TODOS los planes están inactivos
        planes_activos = ot.cliente.planes.filter(activo=True).exists()
        if not planes_activos:
            ot.cliente.estado_activo = False
            ot.cliente.save(update_fields=['estado_activo'])

    elif (
        'reconexion' in nombre_concepto
        or 'reconexión' in nombre_concepto
        or categoria == 'RECONEXION'
    ):
        # Reactivar el plan asociado
        if ot.plan_asociado:
            ot.plan_asociado.activo = True
            ot.plan_asociado.save(update_fields=['activo'])
        # Activar cliente
        ot.cliente.estado_activo = True
        ot.cliente.save(update_fields=['estado_activo'])


@login_required(login_url='admin:login')
def ot_completar(request, pk):
    ot = get_object_or_404(OrdenTecnica, pk=pk)
    ot.completada = True
    ot.fecha_finalizacion = timezone.now()
    ot.save(update_fields=['completada', 'fecha_finalizacion'])
    _aplicar_estado_cliente_por_ot(ot)
    registrar_movimiento(
        ot.cliente,
        'OT completada',
        f"{ot.concepto.nombre} - S/ {ot.monto}",
        icono='fa-check',
        clase='success'
    )
    return redirect(request.META.get('HTTP_REFERER', 'tecnico-lista'))


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def ot_eliminar(request, pk):
    ot = get_object_or_404(OrdenTecnica, pk=pk)
    cliente = ot.cliente
    cliente_id = cliente.pk
    registrar_movimiento(
        cliente,
        'OT eliminada',
        f"{ot.concepto.nombre} - S/ {ot.monto}",
        icono='fa-trash-alt',
        clase='danger'
    )
    ot.delete()
    return redirect('cliente-detalle', pk=cliente_id)


# --- Ajustes Views ---

@login_required(login_url='admin:login')
def ajustes_index(request):
    return render(request, 'billing_app/ajustes/index.html')


@login_required(login_url='admin:login')
def ajustes_importar(request):
    if not is_developer(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    context = {}
    if request.method == 'POST':
        file_obj = request.FILES.get('clientes_xlsx')
        if not file_obj:
            context['import_error'] = 'Selecciona un archivo XLSX.'
        else:
            result = _import_clientes_con_calles_xlsx(file_obj)
            context.update(result)
    return render(request, 'billing_app/ajustes/importar.html', context)


@login_required(login_url='admin:login')
def zona_lista(request):
    distritos = Distrito.objects.all().prefetch_related('sectores__vias')
    return render(
        request,
        'billing_app/ajustes/zona_lista.html',
        {'distritos': distritos}
    )


@login_required(login_url='admin:login')
def servicio_lista(request):
    servicios = Servicio.objects.all().prefetch_related('planes')
    return render(
        request,
        'billing_app/ajustes/servicio_lista.html',
        {'servicios': servicios}
    )


@login_required(login_url='admin:login')
def ot_concepto_lista(request):
    conceptos = OrdenTecnicaConcepto.objects.all().order_by('categoria')
    grouped_data = []
    for cat_code, cat_name in OrdenTecnicaConcepto.CAT_CHOICES:
        grouped_data.append({
            'code': cat_code,
            'name': cat_name,
            'items': conceptos.filter(categoria=cat_code)
        })
    return render(
        request,
        'billing_app/ajustes/ot_lista.html',
        {'grouped_categories': grouped_data}
    )


@login_required(login_url='admin:login')
def serie_lista(request):
    series = SerieCorrelativo.objects.all()
    return render(
        request,
        'billing_app/ajustes/serie_lista.html',
        {'series': series}
    )


@login_required(login_url='admin:login')
def mikrotik_sync(request):
    configs = MikrotikConfig.objects.all().order_by('nombre')
    selected = None
    error = None
    sync_rows = []
    profiles = []
    remote_users = {}
    summary = {
        'ok': 0,
        'missing_mikrotik': 0,
        'missing_local': 0,
        'mismatch': 0
    }

    selected_id = (
        request.POST.get('mikrotik_id') or request.GET.get('mikrotik_id')
    )
    if selected_id:
        selected = MikrotikConfig.objects.filter(pk=selected_id).first()
        if not selected:
            error = 'Mikrotik no encontrado'
        elif request.method == 'POST':
            password = (request.POST.get('password') or '').strip()
            if not password:
                error = 'Clave requerida'
            else:
                try:
                    from routeros_api import RouterOsApiPool
                except ImportError:
                    error = 'Falta instalar routeros-api'
                else:
                    api_pool = None
                    try:
                        api_pool = RouterOsApiPool(
                            selected.ip_host,
                            username=selected.usuario,
                            password=password,
                            port=selected.puerto_api,
                            plaintext_login=True
                        )
                        api = api_pool.get_api()
                        secrets = api.get_resource('/ppp/secret').get()
                        profiles = api.get_resource('/ppp/profile').get()
                        remote_users = {
                            item.get('name'): item for item in secrets
                        }
                    except Exception as exc:
                        error = f'Error de conexion Mikrotik: {exc}'
                    finally:
                        if api_pool:
                            try:
                                api_pool.disconnect()
                            except Exception:
                                pass

    if selected and remote_users:
        local_clientes = (
            Cliente.objects.filter(
                mikrotik_vinculado=selected,
                usuario_pppoe__isnull=False
            )
            .exclude(usuario_pppoe='')
            .prefetch_related('planes__plan__servicio')
        )
        local_lookup = {c.usuario_pppoe: c for c in local_clientes}

        for username, cliente in local_lookup.items():
            remote = remote_users.get(username)
            status = 'OK'
            issues = []
            local_plan = None
            latest_plan = (
                cast(Any, cliente).planes.order_by('-fecha_inicio').first()
            )
            if latest_plan and latest_plan.plan:
                local_plan = latest_plan.plan.nombre
            remote_profile = remote.get('profile') if remote else ''
            local_ip = cliente.ip_asignada or ''
            remote_ip = remote.get('remote-address') if remote else ''

            if not remote:
                status = 'Falta en Mikrotik'
                summary['missing_mikrotik'] += 1
            else:
                if (
                    local_plan and remote_profile
                    and local_plan != remote_profile
                ):
                    issues.append('Perfil distinto')
                if local_ip and remote_ip and local_ip != remote_ip:
                    issues.append('IP distinta')
                if issues:
                    status = 'Desajuste'
                    summary['mismatch'] += 1
                else:
                    summary['ok'] += 1

            sync_rows.append({
                'username': username,
                'cliente': str(cliente),
                'plan_local': local_plan or '',
                'profile_remote': remote_profile or '',
                'ip_local': local_ip,
                'ip_remote': remote_ip or '',
                'status': status,
                'issues': ', '.join(issues)
            })

        for username, remote in remote_users.items():
            if username in local_lookup:
                continue
            summary['missing_local'] += 1
            sync_rows.append({
                'username': username,
                'cliente': '(No registrado)',
                'plan_local': '',
                'profile_remote': remote.get('profile') or '',
                'ip_local': '',
                'ip_remote': remote.get('remote-address') or '',
                'status': 'Falta en sistema',
                'issues': ''
            })

    return render(
        request,
        'billing_app/mikrotik_sync.html',
        {
            'configs': configs,
            'selected': selected,
            'error': error,
            'profiles': profiles,
            'sync_rows': sync_rows,
            'summary': summary
        }
    )


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def mikrotik_status(request, pk):
    config = get_object_or_404(MikrotikConfig, pk=pk)
    password = request.POST.get('password', '')
    if not password:
        return JsonResponse({'ok': False, 'error': 'Clave requerida'})

    try:
        from routeros_api import RouterOsApiPool
    except ImportError:
        return JsonResponse({
            'ok': False,
            'error': 'Falta instalar routeros-api'
        })

    api_pool = None
    try:
        api_pool = RouterOsApiPool(
            config.ip_host,
            username=config.usuario,
            password=password,
            port=config.puerto_api,
            plaintext_login=True
        )
        api = api_pool.get_api()
        resource = api.get_resource('/system/resource').get()
        identity = api.get_resource('/system/identity').get()
        data = resource[0] if resource else {}
        ident = identity[0] if identity else {}
    except Exception as exc:
        return JsonResponse({
            'ok': False,
            'error': f'No conectado: {exc}'
        })
    finally:
        if api_pool:
            try:
                api_pool.disconnect()
            except Exception:
                pass

    config_id = cast(Any, config).id
    connections = request.session.get('mikrotik_connections') or {}
    connections[str(config_id)] = {
        'password': password,
        'verified_at': timezone.now().isoformat()
    }
    request.session['mikrotik_connections'] = connections
    request.session.modified = True

    return JsonResponse({
        'ok': True,
        'data': {
            'identity': ident.get('name', ''),
            'model': data.get('board-name', ''),
            'version': data.get('version', ''),
            'uptime': data.get('uptime', ''),
            'cpu': data.get('cpu', ''),
            'cpu_load': data.get('cpu-load', ''),
            'memory_free': data.get('free-memory', ''),
            'memory_total': data.get('total-memory', '')
        }
    })


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def mikrotik_traffic(request, pk):
    config = get_object_or_404(MikrotikConfig, pk=pk)
    username = (request.POST.get('username') or '').strip()
    profile = (request.POST.get('profile') or '').strip()
    if not username and not profile:
        return JsonResponse({
            'ok': False,
            'error': 'Ingrese usuario PPPoE o perfil'
        })

    connections = request.session.get('mikrotik_connections') or {}
    config_id = cast(Any, config).id
    stored = connections.get(str(config_id))
    if not stored or not stored.get('password'):
        return JsonResponse({
            'ok': False,
            'error': 'Conecte el Mikrotik primero'
        })

    try:
        from routeros_api import RouterOsApiPool
    except ImportError:
        return JsonResponse({
            'ok': False,
            'error': 'Falta instalar routeros-api'
        })

    api_pool = None
    try:
        api_pool = RouterOsApiPool(
            config.ip_host,
            username=config.usuario,
            password=stored['password'],
            port=config.puerto_api,
            plaintext_login=True
        )
        api = api_pool.get_api()
        actives = api.get_resource('/ppp/active').get()
    except Exception as exc:
        return JsonResponse({
            'ok': False,
            'error': f'Error consultando trafico: {exc}'
        })
    finally:
        if api_pool:
            try:
                api_pool.disconnect()
            except Exception:
                pass

    results = []
    for item in actives:
        if username and item.get('name') != username:
            continue
        if profile and item.get('profile') != profile:
            continue
        results.append({
            'name': item.get('name', ''),
            'profile': item.get('profile', ''),
            'address': item.get('address', ''),
            'uptime': item.get('uptime', ''),
            'bytes_in': item.get('bytes-in', ''),
            'bytes_out': item.get('bytes-out', '')
        })

    return JsonResponse({
        'ok': True,
        'rows': results
    })


@login_required(login_url='admin:login')
def mikrotik_lista(request):
    configs = MikrotikConfig.objects.all().prefetch_related('pools')
    connected_ids = set(
        (request.session.get('mikrotik_connections') or {}).keys()
    )
    return render(
        request,
        'billing_app/ajustes/mikrotik_lista.html',
        {'configs': configs, 'connected_ids': connected_ids}
    )


@login_required(login_url='admin:login')
def egreso_concepto_lista(request):
    conceptos = EgresoConcepto.objects.all().order_by('nombre')
    return render(
        request,
        'billing_app/ajustes/egreso_concepto_lista.html',
        {'conceptos': conceptos}
    )


@login_required(login_url='admin:login')
def egreso_lista(request):
    if not (can_manage_caja(request.user) or can_manage_ajustes(request.user)):
        return HttpResponseForbidden('Acceso no autorizado')
    egresos = (
        Egreso.objects.select_related('concepto')
        .order_by('-fecha', '-created_at')
    )
    return render(
        request,
        'billing_app/ajustes/egreso_lista.html',
        {'egresos': egresos}
    )


@login_required(login_url='admin:login')
@require_http_methods(["POST"])
def egreso_eliminar(request, pk):
    if not (can_manage_caja(request.user) or can_manage_ajustes(request.user)):
        return HttpResponseForbidden('Acceso no autorizado')
    egreso = get_object_or_404(Egreso, pk=pk)
    egreso.delete()
    return redirect('egreso-lista')


@login_required(login_url='admin:login')
def roles_lista(request):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    roles = AppRole.objects.all().order_by('nombre')
    asignaciones = (
        UserRole.objects.select_related('user', 'role')
        .order_by('user__username')
    )
    usuarios = get_user_model().objects.all().order_by('username')
    return render(
        request,
        'billing_app/ajustes/roles_lista.html',
        {
            'roles': roles,
            'asignaciones': asignaciones,
            'usuarios': usuarios
        }
    )


# --- Creation & Edit via Side Panel ---

def generic_side_panel_form(
    request, form_class, title, redirect_url, instance=None
):
    if request.method == 'POST':
        form = form_class(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            return redirect(redirect_url)
    else:
        form = form_class(instance=instance)
    return render(
        request,
        'billing_app/ajustes/partial_form.html',
        {'form': form, 'titulo': title}
    )


@login_required(login_url='admin:login')
def distrito_crear(request):
    return generic_side_panel_form(
        request, DistritoForm, 'Nuevo Distrito', 'zona-lista'
    )


@login_required(login_url='admin:login')
def distrito_editar(request, pk):
    obj = get_object_or_404(Distrito, pk=pk)
    return generic_side_panel_form(
        request, DistritoForm, 'Editar Distrito', 'zona-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def distrito_eliminar(request, pk):
    get_object_or_404(Distrito, pk=pk).delete()
    return redirect('zona-lista')


@login_required(login_url='admin:login')
def sector_crear(request):
    return generic_side_panel_form(
        request, SectorForm, 'Nuevo Sector', 'zona-lista'
    )


@login_required(login_url='admin:login')
def sector_editar(request, pk):
    obj = get_object_or_404(Sector, pk=pk)
    return generic_side_panel_form(
        request, SectorForm, 'Editar Sector', 'zona-lista', instance=obj
    )


@login_required(login_url='admin:login')
def sector_eliminar(request, pk):
    get_object_or_404(Sector, pk=pk).delete()
    return redirect('zona-lista')


@login_required(login_url='admin:login')
def via_crear(request):
    return generic_side_panel_form(
        request, ViaForm, 'Nueva Vía', 'zona-lista'
    )


@login_required(login_url='admin:login')
def via_masiva(request, sector_id):
    sector = get_object_or_404(Sector, pk=sector_id)
    ViaFormSet = modelformset_factory(
        Via, fields=('tipo', 'nombre'), extra=10, can_delete=False
    )
    if request.method == 'POST':
        formset = ViaFormSet(
            request.POST, queryset=Via.objects.none()
        )
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                if instance.nombre:
                    instance.sector = sector
                    instance.save()
            return redirect('zona-lista')
    else:
        formset = ViaFormSet(queryset=Via.objects.none())
    return render(
        request,
        'billing_app/ajustes/partial_masivo_vias.html',
        {'formset': formset, 'sector': sector}
    )


@login_required(login_url='admin:login')
def via_editar(request, pk):
    obj = get_object_or_404(Via, pk=pk)
    return generic_side_panel_form(
        request, ViaForm, 'Editar Vía', 'zona-lista', instance=obj
    )


@login_required(login_url='admin:login')
def via_eliminar(request, pk):
    get_object_or_404(Via, pk=pk).delete()
    return redirect('zona-lista')


@login_required(login_url='admin:login')
def sector_masivo(request, distrito_id):
    distrito = get_object_or_404(Distrito, pk=distrito_id)
    SectorFormSet = modelformset_factory(
        Sector, fields=('nombre',), extra=10, can_delete=False
    )
    if request.method == 'POST':
        formset = SectorFormSet(
            request.POST, queryset=Sector.objects.none()
        )
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                if instance.nombre:
                    instance.distrito = distrito
                    instance.save()
            return redirect('zona-lista')
    else:
        formset = SectorFormSet(queryset=Sector.objects.none())
    return render(
        request,
        'billing_app/ajustes/partial_masivo_sectores.html',
        {'formset': formset, 'distrito': distrito}
    )


@login_required(login_url='admin:login')
def servicio_crear(request):
    return generic_side_panel_form(
        request, ServicioForm, 'Nuevo Servicio', 'servicio-lista'
    )


@login_required(login_url='admin:login')
def servicio_editar(request, pk):
    obj = get_object_or_404(Servicio, pk=pk)
    return generic_side_panel_form(
        request, ServicioForm, 'Editar Servicio', 'servicio-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def servicio_eliminar(request, pk):
    get_object_or_404(Servicio, pk=pk).delete()
    return redirect('servicio-lista')


@login_required(login_url='admin:login')
def plan_crear_en_servicio(request, servicio_id):
    servicio = get_object_or_404(Servicio, pk=servicio_id)
    if request.method == 'POST':
        form = PlanForm(request.POST)
        if form.is_valid():
            plan = form.save(commit=False)
            plan.servicio = servicio
            plan.save()
            return redirect('servicio-lista')
    else:
        form = PlanForm(initial={'servicio': servicio})
    return render(
        request,
        'billing_app/ajustes/partial_form.html',
        {'form': form, 'titulo': f'Nuevo Plan para {servicio.nombre}'}
    )


@login_required(login_url='admin:login')
def plan_editar(request, pk):
    obj = get_object_or_404(Plan, pk=pk)
    return generic_side_panel_form(
        request, PlanForm, 'Editar Plan', 'servicio-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def plan_eliminar(request, pk):
    get_object_or_404(Plan, pk=pk).delete()
    return redirect('servicio-lista')


@login_required(login_url='admin:login')
def ot_concepto_crear(request):
    cat = request.GET.get('cat')
    if request.method == 'POST':
        form = OTConceptoForm(request.POST)
        if cat and 'categoria' in form.fields:
            cat_label = dict(OrdenTecnicaConcepto.CAT_CHOICES).get(cat, cat)
            form.fields['categoria'].choices = [(cat, str(cat_label))]
        if form.is_valid():
            form.save()
            return redirect('ot-lista')
    else:
        initial = {}
        if cat:
            initial['categoria'] = cat
        form = OTConceptoForm(initial=initial)
        if cat and 'categoria' in form.fields:
            cat_label = dict(OrdenTecnicaConcepto.CAT_CHOICES).get(cat, cat)
            form.fields['categoria'].choices = [(cat, str(cat_label))]
    return render(
        request,
        'billing_app/ajustes/partial_form.html',
        {'form': form, 'titulo': 'Nuevo Concepto OT'}
    )


@login_required(login_url='admin:login')
def ot_concepto_editar(request, pk):
    obj = get_object_or_404(OrdenTecnicaConcepto, pk=pk)
    return generic_side_panel_form(
        request, OTConceptoForm, 'Editar Concepto OT', 'ot-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def ot_concepto_eliminar(request, pk):
    get_object_or_404(OrdenTecnicaConcepto, pk=pk).delete()
    return redirect('ot-lista')


@login_required(login_url='admin:login')
def serie_crear(request):
    return generic_side_panel_form(
        request, SerieCorrelativoForm, 'Nueva Serie', 'serie-lista'
    )


@login_required(login_url='admin:login')
def serie_editar(request, pk):
    obj = get_object_or_404(SerieCorrelativo, pk=pk)
    return generic_side_panel_form(
        request, SerieCorrelativoForm, 'Editar Serie', 'serie-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def serie_eliminar(request, pk):
    get_object_or_404(SerieCorrelativo, pk=pk).delete()
    return redirect('serie-lista')


@login_required(login_url='admin:login')
def mikrotik_crear(request):
    return generic_side_panel_form(
        request, MikrotikConfigForm, 'Nuevo Mikrotik', 'mikrotik-lista'
    )


@login_required(login_url='admin:login')
def mikrotik_editar(request, pk):
    obj = get_object_or_404(MikrotikConfig, pk=pk)
    return generic_side_panel_form(
        request, MikrotikConfigForm, 'Editar Mikrotik', 'mikrotik-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def mikrotik_eliminar(request, pk):
    get_object_or_404(MikrotikConfig, pk=pk).delete()
    return redirect('mikrotik-lista')


@login_required(login_url='admin:login')
def egreso_concepto_crear(request):
    return generic_side_panel_form(
        request,
        EgresoConceptoForm,
        'Nuevo Concepto de Egreso',
        'egreso-concepto-lista'
    )


@login_required(login_url='admin:login')
def egreso_concepto_editar(request, pk):
    obj = get_object_or_404(EgresoConcepto, pk=pk)
    return generic_side_panel_form(
        request,
        EgresoConceptoForm,
        'Editar Concepto de Egreso',
        'egreso-concepto-lista',
        instance=obj
    )


@login_required(login_url='admin:login')
def egreso_concepto_eliminar(request, pk):
    get_object_or_404(EgresoConcepto, pk=pk).delete()
    return redirect('egreso-concepto-lista')


@login_required(login_url='admin:login')
def role_crear(request):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    return generic_side_panel_form(
        request, AppRoleForm, 'Nuevo Rol', 'roles-lista'
    )


@login_required(login_url='admin:login')
def role_editar(request, pk):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    obj = get_object_or_404(AppRole, pk=pk)
    return generic_side_panel_form(
        request, AppRoleForm, 'Editar Rol', 'roles-lista', instance=obj
    )


@login_required(login_url='admin:login')
def role_eliminar(request, pk):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    get_object_or_404(AppRole, pk=pk).delete()
    return redirect('roles-lista')


@login_required(login_url='admin:login')
def role_asignar(request):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    if request.method == 'POST':
        form = UserRoleForm(request.POST)
        if form.is_valid():
            UserRole.objects.update_or_create(
                user=form.cleaned_data['user'],
                defaults={'role': form.cleaned_data['role']}
            )
            return redirect('roles-lista')
    else:
        form = UserRoleForm()
    return render(
        request,
        'billing_app/ajustes/partial_form.html',
        {'form': form, 'titulo': 'Asignar Rol'}
    )


@login_required(login_url='admin:login')
def usuario_crear(request):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('roles-lista')
    else:
        form = UserCreateForm()
    return render(
        request,
        'billing_app/ajustes/partial_form.html',
        {'form': form, 'titulo': 'Crear Usuario'}
    )


@login_required(login_url='admin:login')
def usuario_editar(request, pk):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    user = get_object_or_404(get_user_model(), pk=pk)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('roles-lista')
    else:
        form = UserEditForm(instance=user)
    return render(
        request,
        'billing_app/ajustes/partial_form.html',
        {'form': form, 'titulo': 'Editar Usuario'}
    )


@login_required(login_url='admin:login')
def usuario_eliminar(request, pk):
    if not can_manage_ajustes(request.user):
        return HttpResponseForbidden('Acceso no autorizado')
    user = get_object_or_404(get_user_model(), pk=pk)
    if request.user.pk == user.pk:
        return HttpResponseForbidden('No puedes eliminar tu usuario activo')
    user.delete()
    return redirect('roles-lista')


@login_required(login_url='admin:login')
def company_settings_edit(request):
    if not (request.user.is_superuser or can_manage_ajustes(request.user)):
        return HttpResponseForbidden('Acceso no autorizado')
    settings_instance = CompanySettings.get_settings()
    if request.method == 'POST':
        form = CompanySettingsForm(
            request.POST, request.FILES, instance=settings_instance
        )
        if form.is_valid():
            form.save()
            return redirect('company-settings')
    else:
        form = CompanySettingsForm(instance=settings_instance)
    return render(
        request,
        'billing_app/ajustes/company_settings.html',
        {'form': form, 'company_settings': settings_instance}
    )
